import openpyxl
import argparse
import json


def read_parameters_from_xl(filename):
    wb = openpyxl.load_workbook(filename, data_only=True)
    parameters = {}
    elements = []
    masses = {}
    for sheetname in wb.get_sheet_names():
        sheet = wb.get_sheet_by_name(sheetname)
        is_after_elements = False
        omega_SR = 0
        gamma_SR = 0
        sigma_SR = 0
        element_col = 0
        proton_col = 0
        mass_col = 0
        for row in range(1, sheet.max_row+1):
            # Parse the header
            if not is_after_elements:
                # The header starts with 'Element'
                if str(sheet.cell(row=row, column=1).value).lower() == "element":
                    # Next loop is after the header
                    is_after_elements = True
                    # Iterate through the header
                    for column in range(1, sheet.max_column+1):
                        # Lower the cell's value and match the names
                        name = str(sheet.cell(row=row, column=column).value).lower()
                        if "element" in name:
                            element_col = column
                        elif "protons" in name:
                            proton_col = column
                        elif "mass" in name:
                            mass_col = column
                        elif "omega_sr" in name:
                            omega_SR = column
                        elif "gamma_sr" in name:
                            gamma_SR = column
                        elif "sigma_sr" in name:
                            sigma_SR = column
                else:
                    # Note the header, so skip
                    continue
            else:
                # Check if all of the needed columndata was found
                total = omega_SR+gamma_SR+sigma_SR+element_col+proton_col+mass_col
                assert total != 0, "Could not find columns"

                # Get the data from the cells
                element = sheet.cell(row=row, column=element_col).value
                protons = sheet.cell(row=row, column=proton_col).value
                mass = sheet.cell(row=row, column=mass_col).value
                omega = sheet.cell(row=row, column=omega_SR).value
                gamma = sheet.cell(row=row, column=gamma_SR).value
                sigma = sheet.cell(row=row, column=sigma_SR).value

                # Create the dicts if missing
                if not element in parameters.keys():
                    parameters[element] = {}

                if not mass in parameters[element].keys():
                    parameters[element][mass] = {}

                # Add elements if missing
                if element not in elements:
                    elements.append(element)

                # Add the element if missing
                if element not in masses.keys():
                    masses[element] = []

                masses[element].append(mass)

                # Store the data
                parameters[element][mass]["epr"] = omega
                parameters[element][mass]["gpr"] = gamma
                parameters[element][mass]["spr"] = sigma
    return parameters, elements, masses


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("xmlfile", help="The Excel file")
    parser.add_argument("jsonfile", help="The JSON file")
    args = parser.parse_args()
    # Get the parameters
    parameters, element, mass = read_parameters_from_xl(args.xmlfile)

    # Load the json file
    with open(args.jsonfile) as rFile:
        data = json.load(rFile)

    # Add the parameters
    data["nested"] = parameters
    data["keywords"]["mass"] = mass
    data["keywords"]["element"] = element

    # Save the amended json file
    with open(args.jsonfile, 'w') as wFile:
        json.dump(data, wFile, indent=4)
